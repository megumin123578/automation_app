import os
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
import os


def save_assignments_to_excel(assignments, out_path, extra_col_names=None):
    wb = Workbook()
    ws = wb.active
    ws.title = "Assignments"

    # ===== Header mặc định =====
    base_headers = ["channel", "directory", "title", "description", "publish_date", "publish_time"]

    # chuẩn hóa extra_col_names
    if isinstance(extra_col_names, str):
        extra_col_names = [extra_col_names]
    elif not extra_col_names:
        extra_col_names = []

    headers = base_headers + extra_col_names
    ws.append(headers)

    # ===== Ghi dữ liệu theo thứ tự row =====
    for r_idx, row in enumerate(assignments, start=2):
        for c_idx, val in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx).value = val

    # ===== In đậm header =====
    for c_idx in range(1, len(headers) + 1):
        ws.cell(row=1, column=c_idx).font = Font(bold=True)

    # ===== Auto width =====
    for col_idx in range(1, len(headers) + 1):
        col_letter = get_column_letter(col_idx)
        max_len = 0
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row,
                                min_col=col_idx, max_col=col_idx):
            val = row[0].value
            if val is None:
                continue
            s = str(val)
            if col_idx == 4:  # cột description
                s = s[:120]
            max_len = max(max_len, len(s))
        ws.column_dimensions[col_letter].width = min(max_len + 2, 80)

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"

    if os.path.exists(out_path):
        os.remove(out_path)
    wb.save(out_path)
    return out_path


def combine_excels(input_dir, output_file, move_folder, get_mp4_filename):
    import glob
    from openpyxl import Workbook, load_workbook
    import os

    # 1) Thu thập file .xlsx
    pattern = os.path.join(input_dir, "*.xlsx")
    files = [f for f in glob.glob(pattern) if os.path.abspath(f) != os.path.abspath(output_file)]
    if not files:
        return 0, []

    # 2) Định nghĩa thứ tự cột target
    BASE = ["channel", "directory", "title", "description", "publish_date", "publish_time"]

    def norm(s):
        return (str(s).strip().lower() if s is not None else "")

    def read_headers(ws):
        return [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]

    # Scan xem có cột monetization trong bất kỳ file nào không
    has_monet = False
    for f in files:
        wb = load_workbook(f)
        ws = wb.active
        headers = [norm(h) for h in read_headers(ws)]
        if "monetization" in headers:
            has_monet = True
            break

    # Thứ tự cột GẮN CHẶT khi gộp:
    # - luôn có move_folder
    # - nếu có monetization ở bất kỳ file nào -> đưa monetization ngay sau move_folder
    ORDERED = BASE + ["move_folder"] + (["monetization"] if has_monet else [])

    # 3) Tạo file output
    wb_out = Workbook()
    ws_out = wb_out.active
    ws_out.title = "Combined"

    # Ghi header theo ORDERED
    for c_idx, name in enumerate(ORDERED, start=1):
        ws_out.cell(row=1, column=c_idx).value = name
    row_out = 2

    processed = []

    for f in files:
        wb = load_workbook(f)
        ws = wb.active

        # Map tên cột -> vị trí cột trong file nguồn
        src_headers_raw = read_headers(ws)
        src_headers = [norm(h) for h in src_headers_raw]
        src_idx = {h: i + 1 for i, h in enumerate(src_headers) if h}

        max_row = ws.max_row
        if max_row < 2:
            processed.append(f)
            continue

        # Với từng dòng dữ liệu nguồn -> ghi ra theo ORDERED
        for r in range(2, max_row + 1):
            # Lấy sẵn các base fields theo tên
            vals_by_name = {}
            for name in BASE:
                c = src_idx.get(name)
                vals_by_name[name] = ws.cell(row=r, column=c).value if c else None

            # move_folder: ưu tiên cột có sẵn, nếu không có -> tính từ directory
            if "move_folder" in src_idx:
                vals_by_name["move_folder"] = ws.cell(row=r, column=src_idx["move_folder"]).value
            else:
                directory_val = vals_by_name.get("directory") or ""
                fn = get_mp4_filename(directory_val) if directory_val else ""
                vals_by_name["move_folder"] = os.path.join(move_folder, fn) if fn else ""

            # monetization: lấy nếu có, còn không thì rỗng
            if has_monet:
                if "monetization" in src_idx:
                    vals_by_name["monetization"] = ws.cell(row=r, column=src_idx["monetization"]).value
                else:
                    vals_by_name["monetization"] = ""

            # Ghi ra đúng thứ tự ORDERED
            for c_idx, name in enumerate(ORDERED, start=1):
                ws_out.cell(row=row_out, column=c_idx).value = vals_by_name.get(name, "")
            row_out += 1

        processed.append(f)

    # Lưu
    os.makedirs(os.path.dirname(output_file) or ".", exist_ok=True)
    if os.path.exists(output_file):
        os.remove(output_file)
    wb_out.save(output_file)

    # Xoá dữ liệu trong file gốc (giữ header)
    for f in processed:
        wb = load_workbook(f)
        ws = wb.active
        if ws.max_row > 1:
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
                for cell in row:
                    cell.value = None
            wb.save(f)

    return len(processed), processed
